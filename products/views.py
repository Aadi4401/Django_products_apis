from django.shortcuts import render
from rest_framework import viewsets, status, mixins
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Product
from .serializers import ProductSerializer
from .permissions import IsAdminOrReadOnly
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone

import openpyxl
from openpyxl.utils import get_column_letter

class ProductViewSet(viewsets.ModelViewSet):
    
    queryset = Product.objects.filter(is_active=True).order_by("-created_on")
    serializer_class = ProductSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = {
        'price': ['gte', 'lte', 'exact'],
        'title': ['exact', 'icontains'],
    }
    search_fields = ['title', 'description']
    ordering_fields = ['created_on', 'updated_on', 'price']
    parser_classes = [MultiPartParser, FormParser]
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.is_active = False
        instance.updated_on = timezone.now()
        instance.save(update_fields=["is_active", "updated_on"])
        return Response(
            {"detail": f"Product {instance.id} soft-deleted."},
            status=status.HTTP_200_OK,
        )


    @action(detail=True, methods=["patch"], url_path="disable", permission_classes=[IsAdminOrReadOnly])
    def disable(self, request, pk=None):
        
        product = self.get_object()
        if not product.is_active:
            return Response({"detail": "Product is already disabled."}, status=status.HTTP_400_BAD_REQUEST)

        product.is_active = False
        product.updated_on = timezone.now()
        product.save(update_fields=["is_active", "updated_on"])
        return Response(
            {"detail": f"Product {product.id} disabled successfully."},
            status=status.HTTP_200_OK,
        )


    @action(detail=False, methods=['post'], url_path='bulk_create', permission_classes=[IsAdminOrReadOnly])
    def bulk_create(self, request):
        
        data = request.data
        if not isinstance(data, list):
            return Response({"detail": "Expected a list"}, status=400)
        serializer = ProductSerializer(data=data, many=True)
        serializer.is_valid(raise_exception=True)
        with transaction.atomic():
            serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='export', permission_classes=[IsAdminOrReadOnly])
    def export(self, request):
        
        include_inactive = request.query_params.get('include_inactive') == 'true'
        qs = Product.objects.all() if include_inactive else Product.objects.filter(is_active=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["id", "title", "description", "price", "discount", "ssn", "is_active", "created_on", "updated_on"]
        ws.append(headers)

        for o in qs:
            ws.append([
                o.id, o.title, o.description, float(o.price), float(o.discount),
                o.ssn, o.is_active,
                o.created_on.strftime("%Y-%m-%d %H:%M:%S"),
                o.updated_on.strftime("%Y-%m-%d %H:%M:%S"),
            ])

        for i, _ in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].auto_size = True

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=products.xlsx'
        wb.save(response)
        return response
